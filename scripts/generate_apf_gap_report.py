#!/usr/bin/env python3
from __future__ import annotations

import argparse
import datetime as dt
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


@dataclass(frozen=True)
class DeltapointFunctionRow:
    row: int
    funcao: str
    tipo: str
    ctl: str | None
    pfb: float
    pfl: float


@dataclass(frozen=True)
class DeltapointResumo:
    sistema: str
    contador: str
    tipo_contagem: str
    total_pf: float


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def load_deltapoint_xlsx(xlsx_path: Path) -> tuple[DeltapointResumo, list[DeltapointFunctionRow]]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    if "Resumo" not in wb.sheetnames or "Funções" not in wb.sheetnames:
        raise ValueError("Planilha inesperada: abas 'Resumo' e 'Funções' nao encontradas.")

    ws_resumo = wb["Resumo"]
    ws_funcoes = wb["Funções"]

    sistema = _cell_str(ws_resumo["D10"].value) or "Portal Candidato"
    contador = _cell_str(ws_resumo["D12"].value) or ""
    tipo_contagem = _cell_str(ws_resumo["D13"].value) or ""
    total_pf = float(ws_resumo["D15"].value or 0)

    rows: list[DeltapointFunctionRow] = []
    for r in range(3, ws_funcoes.max_row + 1):
        funcao = _cell_str(ws_funcoes.cell(r, 1).value)
        tipo = _cell_str(ws_funcoes.cell(r, 2).value)
        if not funcao and not tipo:
            continue
        if not funcao or not tipo:
            continue

        ctl = _cell_str(ws_funcoes.cell(r, 10).value) or None  # J: ctl (ex: SEA, EEA, ALIL)
        pfb = float(ws_funcoes.cell(r, 16).value or 0)  # P: PFB
        pfl = float(ws_funcoes.cell(r, 17).value or 0)  # Q: PFL

        rows.append(
            DeltapointFunctionRow(
                row=r,
                funcao=funcao,
                tipo=tipo,
                ctl=ctl,
                pfb=pfb,
                pfl=pfl,
            )
        )

    resumo = DeltapointResumo(
        sistema=sistema,
        contador=contador,
        tipo_contagem=tipo_contagem,
        total_pf=total_pf,
    )
    return resumo, rows


def count_api_endpoints(controllers_dir: Path) -> list[tuple[str, int]]:
    pattern = re.compile(r"\[Http(Get|Post|Put|Delete|Patch)\b")
    items: list[tuple[str, int]] = []
    for p in sorted(controllers_dir.glob("*Controller.cs")):
        if p.name == "BaseController.cs":
            continue
        text = p.read_text(encoding="utf-8")
        cnt = len(pattern.findall(text))
        items.append((p.name, cnt))
    # Desc by endpoints
    items.sort(key=lambda x: x[1], reverse=True)
    return items


def build_pdf(
    output_pdf: Path,
    deltapoint_xlsx_name: str,
    deltapoint_resumo: DeltapointResumo,
    deltapoint_rows: list[DeltapointFunctionRow],
    controller_counts: list[tuple[str, int]],
) -> None:
    styles = getSampleStyleSheet()
    styles.add(
        ParagraphStyle(
            name="H1",
            parent=styles["Heading1"],
            fontName="Helvetica-Bold",
            fontSize=14,
            spaceAfter=8,
        )
    )
    styles.add(
        ParagraphStyle(
            name="H2",
            parent=styles["Heading2"],
            fontName="Helvetica-Bold",
            fontSize=12,
            spaceAfter=6,
        )
    )
    styles.add(
        ParagraphStyle(
            name="Small",
            parent=styles["BodyText"],
            fontName="Helvetica",
            fontSize=9.5,
            leading=12,
        )
    )

    analysis_date = dt.date.today().strftime("%Y-%m-%d")

    def P(text: str, style_name: str = "Normal") -> Paragraph:
        # Use Paragraph to ensure proper wrapping inside Table cells.
        return Paragraph(text, styles[style_name])

    # Deltapoint breakdown
    tipo_counts = Counter(r.tipo for r in deltapoint_rows)
    pfb_total = sum(r.pfb for r in deltapoint_rows)
    pfl_total = sum(r.pfl for r in deltapoint_rows)

    # API endpoints breakdown
    total_controllers = len(controller_counts)
    total_endpoints = sum(cnt for _, cnt in controller_counts)

    # Build document
    doc = SimpleDocTemplate(
        str(output_pdf),
        pagesize=A4,
        leftMargin=2.0 * cm,
        rightMargin=2.0 * cm,
        topMargin=2.0 * cm,
        bottomMargin=2.0 * cm,
        title="Relatorio Tecnico - Gap APF",
        author="Equipe Tecnica",
    )

    story: list[Any] = []

    story.append(Paragraph("Relatorio Tecnico - Analise de Gap em APF", styles["Title"]))
    story.append(Paragraph("Projeto: CAU Sistema Eleitoral (Migrado)", styles["Normal"]))
    story.append(Paragraph(f"Data do relatorio: {analysis_date}", styles["Normal"]))
    story.append(Spacer(1, 10))

    story.append(Paragraph("1. Objetivo", styles["H1"]))
    story.append(
        Paragraph(
            "Este documento consolida a comparacao entre a contagem estimada da Deltapoint "
            "para o Portal do Candidato e as evidencias encontradas no codigo-fonte do projeto, "
            "explicando as principais causas do gap (diferenca) de Pontos de Funcao (APF) e propondo "
            "um caminho de reconciliacao.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    story.append(Paragraph("2. Fontes Consideradas", styles["H1"]))
    story.append(
        Paragraph(
            f"- Planilha: {deltapoint_xlsx_name} (Deltapoint).<br/>"
            "- Codigo-fonte do repositorio CAU Sistema Eleitoral Migrado (.NET + React).",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    story.append(Paragraph("3. Resumo Executivo", styles["H1"]))

    resumo_data = [
        ["Item", "Valor", "Observacao"],
        [
            P("Deltapoint (Portal Candidato)", "Small"),
            P(f"{deltapoint_resumo.total_pf:.0f} PF", "Small"),
            P(f"{len(deltapoint_rows)} funcoes identificadas na planilha (contagem estimada).", "Small"),
        ],
        [
            P("Distribuicao Deltapoint", "Small"),
            P(
                f"{tipo_counts.get('ALI', 0)} ALI, {tipo_counts.get('EE', 0)} EE, {tipo_counts.get('CE', 0)} CE, {tipo_counts.get('SE', 0)} SE",
                "Small",
            ),
            P(f"Total PFB={pfb_total:.0f}, PFL={pfl_total:.0f}.", "Small"),
        ],
        [
            P("API (evidencia em codigo)", "Small"),
            P(f"{total_endpoints} endpoints HTTP", "Small"),
            P(f"{total_controllers} controllers funcionais (exclui BaseController).", "Small"),
        ],
        [
            P("Nosso baseline (sistema completo)", "Small"),
            P("2.474 PF nao ajustados (2.870 ajustados, VAF 1,16)", "Small"),
            P(
                "Contagem detalhada baseada em analise estatica do codigo e estimativas de DET/RET/FTR.",
                "Small",
            ),
        ],
        [
            P("Nossa reconciliacao (IFPUG: ALI agrupado)", "Small"),
            P("1.111 PF nao ajustados (1.289 ajustados)", "Small"),
            P("Agrupa ALIs logicamente e consolida transacoes por processo elementar.", "Small"),
        ],
    ]

    t = Table(resumo_data, colWidths=[5.2 * cm, 5.2 * cm, 6.4 * cm])
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, 0), 10),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 1), (-1, -1), 9.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(t)
    story.append(Spacer(1, 10))

    story.append(Paragraph("Interpretacao rapida:", styles["H2"]))
    story.append(
        Paragraph(
            "A planilha da Deltapoint analisada refere-se especificamente ao Portal do Candidato "
            "e contabiliza um subconjunto pequeno de funcionalidades (26 funcoes). "
            "Os numeros do sistema completo (incluindo modulos administrativos, votacao, apuracao, "
            "denuncias, impugnacoes, relatorios etc.) sao naturalmente muito maiores. "
            "O gap costuma ser explicado por diferencas de escopo e granularidade de transacoes, "
            "alem de decisao de agrupamento de ALIs segundo as regras IFPUG.",
            styles["Normal"],
        )
    )

    story.append(PageBreak())

    story.append(Paragraph("4. O Que a Deltapoint Contou (Portal Candidato)", styles["H1"]))
    story.append(
        Paragraph(
            f"Sistema na planilha: <b>{deltapoint_resumo.sistema}</b><br/>"
            f"Contador: <b>{deltapoint_resumo.contador or 'nao informado'}</b><br/>"
            f"Tipo de contagem: <b>{deltapoint_resumo.tipo_contagem or 'nao informado'}</b>",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    story.append(Paragraph("4.1 Distribuicao por Tipo de Funcao", styles["H2"]))
    dist_data = [
        ["Tipo", "Qtde", "Peso tipico (IFPUG)", "PF na planilha (PFB)"],
    ]

    # Valores observados no arquivo: ALI simples (7), EE medio (4), CE medio (4), SE medio (5)
    weights = {"ALI": "7", "AIE": "5", "EE": "4", "CE": "4", "SE": "5"}
    pf_by_type = {"ALI": 0.0, "AIE": 0.0, "EE": 0.0, "CE": 0.0, "SE": 0.0}
    for r in deltapoint_rows:
        pf_by_type[r.tipo] = pf_by_type.get(r.tipo, 0.0) + r.pfb

    for tipo in ["ALI", "AIE", "EE", "CE", "SE"]:
        if tipo_counts.get(tipo, 0) == 0 and pf_by_type.get(tipo, 0.0) == 0:
            continue
        dist_data.append([tipo, str(tipo_counts.get(tipo, 0)), weights.get(tipo, "-"), f"{pf_by_type.get(tipo, 0.0):.0f}"])

    dist_data.append(["Total", str(len(deltapoint_rows)), "-", f"{pfb_total:.0f}"])

    t2 = Table(dist_data, colWidths=[2.2 * cm, 2.0 * cm, 4.2 * cm, 8.4 * cm])
    t2.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8EEF7")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(t2)
    story.append(Spacer(1, 10))

    story.append(Paragraph("4.2 Lista de Funcoes (26 itens)", styles["H2"]))
    story.append(
        Paragraph(
            "Abaixo estao as funcoes exatamente como aparecem na planilha analisada (Portal Candidato).",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))

    funcs_table = [["#", "Tipo", "CTL", "Funcao", "PF (PFB)"]]
    for idx, r in enumerate(deltapoint_rows, start=1):
        funcs_table.append([str(idx), r.tipo, r.ctl or "-", P(r.funcao, "Small"), f"{r.pfb:.0f}"])

    t3 = Table(funcs_table, colWidths=[0.9 * cm, 1.3 * cm, 1.6 * cm, 10.5 * cm, 1.7 * cm])
    t3.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                ("LEFTPADDING", (0, 0), (-1, -1), 5),
                ("RIGHTPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(t3)

    story.append(PageBreak())

    story.append(Paragraph("5. Evidencias no Codigo (API)", styles["H1"]))
    story.append(
        Paragraph(
            "A seguir esta a contagem de endpoints HTTP por controller no codigo-fonte (.NET). "
            "A contagem considera metodos com atributos [HttpGet], [HttpPost], [HttpPut], [HttpDelete] e [HttpPatch].",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 8))

    api_table = [["Controller", "Endpoints"]]
    for name, cnt in controller_counts:
        api_table.append([name.replace("Controller.cs", "Controller"), str(cnt)])
    api_table.append(["Total", str(total_endpoints)])

    t4 = Table(api_table, colWidths=[10.0 * cm, 4.8 * cm])
    t4.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                ("FONTSIZE", (0, 0), (-1, -1), 9.5),
                ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.whitesmoke, colors.white]),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#E8EEF7")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("LEFTPADDING", (0, 0), (-1, -1), 6),
                ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(t4)
    story.append(Spacer(1, 10))

    story.append(Paragraph("6. Principais Causas do Gap", styles["H1"]))
    story.append(
        Paragraph(
            "Os gaps de APF normalmente decorrem de decisoes metodologicas. Os fatores abaixo explicam, "
            "de forma objetiva, por que uma contagem de portal isolado pode ficar muito abaixo de uma contagem "
            "do sistema completo:",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "<b>6.1 Escopo</b><br/>"
            "- A planilha analisada cobre somente o Portal do Candidato.<br/>"
            "- O sistema completo inclui modulos administrativos, votacao, apuracao, relatorios, auditoria, "
            "configuracoes, notificacoes, denuncias, impugnacoes e workflows associados.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "<b>6.2 Granularidade das transacoes</b><br/>"
            "- Uma contagem pode agrupar varios endpoints em um unico processo elementar (visao do usuario).<br/>"
            "- Outra contagem pode contar cada endpoint separadamente, inflando EEs/CEs/SEs.<br/>"
            "- Nossa reconciliacao (1.111 PF nao ajustados) tenta alinhar com processo elementar, sem perder escopo.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "<b>6.3 Agrupamento de ALIs</b><br/>"
            "- IFPUG define ALI como um grupo logico de dados reconhecivel pelo usuario, nao como tabela isolada.<br/>"
            "- Contar 156 tabelas como 156 ALIs gera valores maiores do que agrupar em ~16 ALIs logicos.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "<b>6.4 Estimativa vs detalhamento</b><br/>"
            "- A planilha da Deltapoint esta como contagem estimada e, no arquivo analisado, "
            "nao apresenta DET/FTR preenchidos para cada funcao.<br/>"
            "- Isso tende a padronizar pesos e reduzir variacoes de complexidade.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 10))

    story.append(Paragraph("7. Proposta de Reconciliacao (para fechar um numero consensual)", styles["H1"]))
    story.append(
        Paragraph(
            "Para chegar a um numero que ambas as partes considerem auditavel, sugerimos uma sessao conjunta "
            "de reconciliacao com os passos abaixo:",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        Paragraph(
            "1) Definir fronteira e escopo (o que entra e o que sai).<br/>"
            "2) Listar processos elementares por modulo (visao do usuario) e mapear para evidencias: telas, "
            "endpoints, historias/requisitos.<br/>"
            "3) Consolidar duplicidades (consultas implicitas, variacoes de filtro) e validar DET/FTR quando necessario.<br/>"
            "4) Publicar planilha final com rastreabilidade por funcao.",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 10))

    story.append(
        Paragraph(
            "<b>Observacao importante:</b> este relatorio nao afirma que a contagem de uma das partes esta 'errada'. "
            "Ele demonstra que, com o arquivo de Portal Candidato (124 PF) e com as evidencias do codigo, "
            "as diferencas sao explicaveis por escopo e granularidade. "
            "Para validar um numero do sistema completo (ex.: 590 PF), e necessario ter a planilha completa "
            "e/ou o detalhamento de funcoes utilizadas nessa totalizacao.",
            styles["Small"],
        )
    )

    doc.build(story)


def main() -> int:
    parser = argparse.ArgumentParser(description="Gera PDF de analise de gap APF (Deltapoint vs codigo).")
    parser.add_argument(
        "--deltapoint-xlsx",
        type=Path,
        default=Path.home() / "Downloads" / "Deltapoint_Contagem_Estimada_PortalCandidato-202602.xlsx",
        help="Caminho para o XLSX da Deltapoint.",
    )
    parser.add_argument(
        "--controllers-dir",
        type=Path,
        default=Path("apps/api/CAU.Eleitoral.Api/Controllers"),
        help="Diretorio com controllers da API.",
    )
    parser.add_argument(
        "--output-pdf",
        type=Path,
        default=Path("output/pdf/relatorio-gap-apf-deltapoint-vs-codigo.pdf"),
        help="Caminho de saida do PDF.",
    )
    args = parser.parse_args()

    xlsx_path: Path = args.deltapoint_xlsx
    controllers_dir: Path = args.controllers_dir
    output_pdf: Path = args.output_pdf

    if not xlsx_path.exists():
        raise SystemExit(f"Arquivo XLSX nao encontrado: {xlsx_path}")
    if not controllers_dir.exists():
        raise SystemExit(f"Diretorio de controllers nao encontrado: {controllers_dir}")

    output_pdf.parent.mkdir(parents=True, exist_ok=True)

    resumo, rows = load_deltapoint_xlsx(xlsx_path)
    controller_counts = count_api_endpoints(controllers_dir)

    build_pdf(
        output_pdf=output_pdf,
        deltapoint_xlsx_name=xlsx_path.name,
        deltapoint_resumo=resumo,
        deltapoint_rows=rows,
        controller_counts=controller_counts,
    )

    print(str(output_pdf))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
